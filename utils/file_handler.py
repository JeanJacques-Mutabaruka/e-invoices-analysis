import streamlit as st
import os
import boto3
import string
import random
import pandas as pd
import locale
import openpyxl
import xlrd
import xlwt
import pyxlsb
import ast
import re
from io import BytesIO
from openpyxl.styles import Color, PatternFill, Font, Border, Alignment
from openpyxl.utils.exceptions import IllegalCharacterError
from datetime import datetime, timedelta

from typing import List, Dict, Tuple, Optional
# NEW:
# Get project root (2 levels up from utils/)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
str_my_MEDIA_ROOT = MEDIA_ROOT = os.path.join(BASE_DIR, 'Findap_mediafiles')

str_my_AWS_STORAGE_BUCKET_NAME = st.secrets["AWS_STORAGE_BUCKET_NAME"]
str_my_AWS_ACCESS_KEY_ID = st.secrets["AWS_ACCESS_KEY_ID"]
str_my_AWS_SECRET_ACCESS_KEY = st.secrets["AWS_SECRET_ACCESS_KEY"]

# ============================================================================
# UTILITY FUNCTIONS TO ADD TO cls_Customfiles_Filetypehandler CLASS
# ============================================================================

def add_date_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add YEAR, MONTH, DAY, YEAR-MONTH columns from TRANSACTION DATE
    Case-insensitive column search
    
    Args:
        df: DataFrame with TRANSACTION DATE column
        
    Returns:
        DataFrame with added date columns
    """
    # Find TRANSACTION DATE column (case insensitive)
    date_col = None
    for col in df.columns:
        if col.upper().strip() == 'TRANSACTION DATE':
            date_col = col
            break
    
    if date_col is None:
        print("⚠️ Warning: TRANSACTION DATE column not found")
        return df
    
    try:
        # Convert to datetime
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        
        # Add new columns
        df['YEAR'] = df[date_col].dt.strftime('%Y')
        df['MONTH'] = df[date_col].dt.strftime('%b')  # MMM format (Jan, Feb, etc.)
        df['DAY'] = df[date_col].dt.strftime('%d')
        df['YEAR-MONTH'] = df[date_col].dt.strftime('%Y-%m')
        
        print(f"✅ Added date columns: YEAR, MONTH, DAY, YEAR-MONTH")
        
    except Exception as e:
        print(f"⚠️ Warning: Could not add date columns: {e}")
    
    return df

def check_duplicate_status_column(df: pd.DataFrame) -> Tuple[bool, pd.DataFrame]:
    """
    Check if Duplicate Status column exists (case insensitive)
    
    Args:
        df: DataFrame to check
        
    Returns:
        Tuple of (column_exists: bool, dataframe with standardized column name)
    """
    # Find Duplicate Status column (case insensitive)
    dup_status_col = None
    for col in df.columns:
        if col.upper().strip() == 'DUPLICATE STATUS':
            dup_status_col = col
            break
    
    if dup_status_col is None:
        return False, df
    
    # Standardize column name if it exists but with different case
    if dup_status_col != 'Duplicate Status':
        df = df.rename(columns={dup_status_col: 'Duplicate Status'})
    
    return True, df

def get_numeric_columns(df: pd.DataFrame, exclude_patterns: List[str] = None) -> List[str]:
    """
    Get list of numeric columns excluding certain patterns
    
    Args:
        df: DataFrame to analyze
        exclude_patterns: List of patterns to exclude (default: JJ's exclusion list)
        
    Returns:
        List of numeric column names
    """
    if exclude_patterns is None:
        # Use JJ's specified exclusion list
        exclude_patterns = ['ID', 'LINE NUMBER', 'INVOICE NUMBER', 'RECEIPT NUMBER', 
                          'YEAR', 'MONTH', 'DAY', 'LINE_NUMBER']
    
    numeric_cols = []
    for col in df.columns:
        # Check if column is numeric
        if pd.api.types.is_numeric_dtype(df[col]):
            # Check if column name contains any exclusion pattern
            col_upper = col.upper()
            should_exclude = False
            
            for pattern in exclude_patterns:
                if pattern.upper() in col_upper:
                    should_exclude = True
                    break
            
            if not should_exclude:
                numeric_cols.append(col)
    
    return numeric_cols

def format_date_for_display(date_val) -> str:
    """
    Format date in dd-mmm-yyyy format
    
    Args:
        date_val: Date value (datetime, string, or timestamp)
        
    Returns:
        Formatted date string
    """
    try:
        if pd.isna(date_val):
            return "N/A"
        
        if isinstance(date_val, str):
            date_val = pd.to_datetime(date_val, errors='coerce')
        
        if pd.isna(date_val):
            return "N/A"
        
        return date_val.strftime('%d-%b-%Y')
    
    except:
        return "N/A"

def get_date_range(df: pd.DataFrame) -> Tuple[str, str]:
    """
    Get date range from TRANSACTION DATE column
    
    Args:
        df: DataFrame with TRANSACTION DATE
        
    Returns:
        Tuple of (start_date, end_date) formatted as dd-mmm-yyyy
    """
    # Find TRANSACTION DATE column (case insensitive)
    date_col = None
    for col in df.columns:
        if col.upper().strip() == 'TRANSACTION DATE':
            date_col = col
            break
    
    if date_col is None:
        return "N/A", "N/A"
    
    try:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        valid_dates = df[date_col].dropna()
        
        if valid_dates.empty:
            return "N/A", "N/A"
        
        start_date = format_date_for_display(valid_dates.min())
        end_date = format_date_for_display(valid_dates.max())
        
        return start_date, end_date
    
    except:
        return "N/A", "N/A"


class ComparisonRulesManager:
    """
    Manages comparison rules loaded from FINDAP_Filetypes_Parameters.xlsx
    Note: For Phase 1, comparison rules between categories are NOT implemented yet.
    This class is prepared for future use when cross-category comparisons are enabled.
    """
    
    df_comparison_rules = None
    df_duplicate_config = None
    df_numeric_config = None
    df_auto_analysis_config = None
    rules_loaded = False
    
    @classmethod
    def load_comparison_rules(cls):
        """Load all comparison-related configurations from parameter file"""
        if cls.rules_loaded:
            return
        
        try:
            # Note: This will be used in future phases for cross-category comparisons
            # For now, we're focusing on single category analysis
            
            # Import here to avoid circular imports
            from utils.file_handler import cls_Customfiles_Filetypehandler
            
            # Load parameters if not already loaded
            cls_Customfiles_Filetypehandler.load_parameters()
            
            param_file = cls_Customfiles_Filetypehandler.str_local_filepath
            
            # Load comparison rules (for future use)
            try:
                cls.df_comparison_rules = pd.read_excel(
                    param_file, 
                    sheet_name='COMPARISON', 
                    na_filter=False
                )
                # Filter only enabled rules
                cls.df_comparison_rules = cls.df_comparison_rules[
                    cls.df_comparison_rules['ENABLED'].astype(str).str.upper() == 'TRUE'
                ]
                # Sort by priority
                cls.df_comparison_rules = cls.df_comparison_rules.sort_values('PRIORITY')
                
            except Exception as e:
                print(f"Info: COMPARISON sheet not found (will be used in future phases): {e}")
                cls.df_comparison_rules = pd.DataFrame()
            
            # Load numeric fields config
            try:
                cls.df_numeric_config = pd.read_excel(
                    param_file,
                    sheet_name='NUMERIC_FIELDS_CONFIG',
                    na_filter=False
                )
            except Exception as e:
                print(f"Info: NUMERIC_FIELDS_CONFIG sheet not found, using defaults: {e}")
                cls.df_numeric_config = pd.DataFrame()
            
            # Load auto-analysis config
            try:
                cls.df_auto_analysis_config = pd.read_excel(
                    param_file,
                    sheet_name='AUTO_ANALYSIS_CONFIG',
                    na_filter=False
                )
            except Exception as e:
                print(f"Info: AUTO_ANALYSIS_CONFIG sheet not found, using defaults: {e}")
                cls.df_auto_analysis_config = pd.DataFrame()
            
            cls.rules_loaded = True
            print(f"✅ Configuration loaded successfully")
            
        except Exception as e:
            print(f"Warning: Error loading analysis configuration: {e}")
            cls.df_comparison_rules = pd.DataFrame()
    
    @classmethod
    def get_numeric_field_config(cls) -> Tuple[List[str], List[str]]:
        """
        Get numeric field configuration
        
        Returns:
            Tuple of (include_patterns, exclude_fields)
        """
        cls.load_comparison_rules()
        
        if cls.df_numeric_config.empty:
            # Default configuration as specified by JJ
            include = ['AMOUNT', 'VAT', 'TOTAL', 'TAXABLE', 'TAX', 'GROSS', 'NET']
            exclude = ['ID', 'LINE NUMBER', 'INVOICE NUMBER', 'RECEIPT NUMBER', 
                      'YEAR', 'MONTH', 'DAY', 'LINE_NUMBER']
            return include, exclude
        
        include = cls.df_numeric_config[
            cls.df_numeric_config['ACTION'] == 'INCLUDE'
        ]['FIELD_PATTERN'].tolist()
        
        exclude = cls.df_numeric_config[
            cls.df_numeric_config['ACTION'] == 'EXCLUDE'
        ]['FIELD_PATTERN'].tolist()
        
        return include, exclude
    
    @classmethod
    def should_run_auto_analysis(cls) -> bool:
        """Check if automatic analysis is enabled"""
        cls.load_comparison_rules()
        
        if cls.df_auto_analysis_config.empty:
            return True  # Default: enabled
        
        # Get first row (global config)
        config = cls.df_auto_analysis_config.iloc[0]
        return str(config.get('ENABLED', 'TRUE')).upper() == 'TRUE'
    
    @classmethod
    def get_auto_analysis_depth(cls) -> str:
        """Get configured analysis depth"""
        cls.load_comparison_rules()
        
        if cls.df_auto_analysis_config.empty:
            return 'standard'
        
        config = cls.df_auto_analysis_config.iloc[0]
        return config.get('ANALYSIS_DEPTH', 'standard').lower()

class cls_Aws_ParamfileHandler:
	"""
	Handles AWS S3 interactions for downloading parameter files needed for processing.
	"""

	# def __init__(self, obj_logger):
	def __init__(self):
		"""
		Initializes the AWS handler with a logger instance.
		"""
		# self.obj_logger = obj_logger  # Centralized logger
		self.obj_s3 = self.fn_initialize_s3_client()
		self.str_bucket_name = str_my_AWS_STORAGE_BUCKET_NAME

	def fn_initialize_s3_client(self):
		"""Initializes an S3 client """
		try:
			obj_session = boto3.Session(
				aws_access_key_id = st.secrets["AWS_ACCESS_KEY_ID"],
				aws_secret_access_key = st.secrets["AWS_SECRET_ACCESS_KEY"]
				# aws_access_key_id = str_my_AWS_ACCESS_KEY_ID,
				# aws_secret_access_key = str_my_AWS_SECRET_ACCESS_KEY
			)
			print("S3 client initialized successfully.",)
			return obj_session.resource('s3')
		except Exception as var_error:
			print(f"ERROR: Failed to initialize S3 client...", )
			return None

	def fn_download_paramfile(self, str_aws_folder: str, str_aws_filename: str, str_local_folder: str) -> bool:

		self.__init__
		if self.obj_s3 is None:
			# self.obj_logger.fn_log("S3 client is not initialized. Cannot download parameter file.", "ERROR")
			print("S3 client is not initialized. Cannot download parameter file.", )
			return False

		str_local_filepath = os.path.join(str_local_folder, str_aws_filename)
		str_s3_key = os.path.join(str_aws_folder, str_aws_filename)

		# Check if file already exists locally
		if os.path.isfile(str_local_filepath):
			# self.obj_logger.fn_log(f"INFO: Parameter file '{str_local_filepath}' already exists. Skipping download.", "INFO")
			print(f"INFO: Parameter file '{str_local_filepath}' already exists. Skipping download.",)
			return True  # No need to re-download

		try:
			# self.obj_logger.fn_log(f"Downloading {str_s3_key} from S3...", "INFO")
			self.obj_s3.meta.client.download_file(self.str_bucket_name, str_s3_key, str_local_filepath)
			# self.obj_logger.fn_log(f"SUCCESS: Parameter file downloaded to {str_local_filepath}", "INFO")
			print(f"SUCCESS: Parameter file downloaded to {str_local_filepath}", )
			return True
		except Exception as var_error:
			print(f"ERROR: Failed to download...", )
			return False
	
	def get_fromaws_findap_filetypesparams(str_Awsfolder,str_Awsfilename,str_destination_filepath):
		session = boto3.Session(
			aws_access_key_id = st.secrets["AWS_ACCESS_KEY_ID"], 
			aws_secret_access_key = st.secrets["AWS_SECRET_ACCESS_KEY"], 
			# aws_access_key_id = str_my_AWS_ACCESS_KEY_ID, 
			# aws_secret_access_key = str_my_AWS_SECRET_ACCESS_KEY, 
		)
		# Get Findap file with Parameters for files types from AWS S3 bucket.
		s3 = session.resource('s3')
		bucket_name = str_my_AWS_STORAGE_BUCKET_NAME
		str_fullname = str_Awsfolder + str_Awsfilename
		s3.meta.client.download_file(bucket_name,str_fullname,str_destination_filepath)

class cls_Customfiles_Filetypehandler:    
	# Class-level cache for parameters
	dic_filesparams_dfs = None
	df_Original_headers = None
	df_Findap_newheaders = None
	df_Categoriesparams = None
	df_Fields_formats = None
	lst_dates_regex = None
	df_Findap_GlobalAnalysis_headers = None
	df_Check_duplicates_params = None
	df_finstatment_map = pd.DataFrame()
	bln_parameters_loaded = False  # Flag to track if parameters were loaded

	str_param_folder = "FINDAP_FILES PARAMETERS/"
	str_param_filename = "FINDAP_Filetypes_Parameters.xlsx"
	str_local_filepath = os.path.join(str_my_MEDIA_ROOT, str_param_filename)

	# def __init__(self, obj_logger, obj_aws_param_handler):
	def __init__(self, obj_aws_param_handler):
		"""
		Initialize the file handler with AWS and Logger instances.
		"""
		# self.obj_logger = obj_logger
		self.obj_aws_param_handler = obj_aws_param_handler
		self.str_param_folder = "FINDAP_FILES PARAMETERS/"
		self.str_param_filename = "FINDAP_Filetypes_Parameters.xlsx"
		self.str_local_filepath = os.path.join(str_my_MEDIA_ROOT, self.str_param_filename)

		# Ensure parameter file is downloaded ONCE during initialization
		self.fn_ensure_paramfile()
		self.load_parameters()

	def fn_ensure_paramfile(self):
		"""
		Ensures the AWS parameter file is available locally before proceeding.
		"""
		bln_download_success = self.obj_aws_param_handler.fn_download_paramfile(
			cls_Customfiles_Filetypehandler.str_param_folder,
			cls_Customfiles_Filetypehandler.str_param_filename,
			str_my_MEDIA_ROOT
		)
		if not bln_download_success:
			# self.obj_logger.fn_log("ERROR: Parameter file could not be retrieved. Processing may fail.", "ERROR")
			print("ERROR: Parameter file could not be retrieved. Processing may fail.", )

	@classmethod
	#def load_parameters(cls,obj_logger):
	def load_parameters(cls):
		"""
		Load shared resources once and cache them in class variables.
		"""
		if not cls.bln_parameters_loaded:
			try:
				# Check if the directory exists, and create it if it doesn't
				if not os.path.exists(os.path.normpath(str_my_MEDIA_ROOT)):
					os.makedirs(os.path.normpath(str_my_MEDIA_ROOT), exist_ok=True)  # Creates all intermediate directories if needed
					# print('{}: MEDIA ROOT folder has been created : {}'.format(datetime.now(),os.path.normpath(str_my_MEDIA_ROOT)))
				
				str_local_filepath = os.path.normpath(os.path.join(str_my_MEDIA_ROOT, cls.str_param_filename))
				cls_Aws_ParamfileHandler.get_fromaws_findap_filetypesparams(cls_Customfiles_Filetypehandler.str_param_folder,
												cls_Customfiles_Filetypehandler.str_param_filename,
												str_local_filepath
												)
				cls.dic_filesparams_dfs = pd.read_excel(str_local_filepath, sheet_name=None, na_filter=False, engine="openpyxl")

				cls.df_Original_headers = cls.dic_filesparams_dfs["FileHeaders"]
				cls.df_Findap_newheaders = cls.dic_filesparams_dfs["FindapHeaders"]
				cls.df_Categoriesparams = cls.dic_filesparams_dfs["Categories_sheetnames"]
				cls.df_Fields_formats = cls.dic_filesparams_dfs["Dataformat"]
				cls.df_Dates_regex= cls.dic_filesparams_dfs["Dates formats"]
				cls.lst_dates_regex= cls.df_Dates_regex["Dates formats"].tolist()
				cls.df_Check_duplicates_params= cls.dic_filesparams_dfs["Check_duplicates"]
				cls.df_Findap_GlobalAnalysis_headers= cls.dic_filesparams_dfs["Findap_GlobalAnalysis"]
				cls.df_Findap_Grouping_WHT_headers= cls.dic_filesparams_dfs["Findap_Grouping_WHT"]
				cls.df_Findap_Cashflowdata_headers= cls.dic_filesparams_dfs["Findap_Cashflowdata"]
				cls.df_Findap_RRAdsccombined_headers= cls.dic_filesparams_dfs["Findap_RRAdsccombined"]
				cls.df_finstatment_map= cls.dic_filesparams_dfs["Fin Statements map"]
				cls.df_TB_params= cls.dic_filesparams_dfs["Trial Balance parameters"]
				
				cls.bln_parameters_loaded = True  # Set flag to True

			except Exception as var_error:
				print(f"ERROR: Failed to load parameters...")


	@classmethod
	def fn_get_Uploadedfile_Sheetscategories(cls, str_Folderpath, obj_file, var_mysheet=None, int_headerrow=None):
		"""Handle file and sheet category extraction."""
		cls.load_parameters()  # Ensure shared resources are loaded		
		dic_myfile_categories = {}
		try:
			dic_myfile_dfs = cls.fn_get_file_as_dicdataframes(str(obj_file), None, )

			def fn_find_matching_category(df_mydf, max_iterations=25):
				"""Find matching category for headers in the given dataframe."""
				max_iterations = min(len(df_mydf), max_iterations)
				# print('{} : CATEGORIES = {}'.format(datetime.now(),list(cls.df_Original_headers.columns))) 
				
				for int_myheaderrow in range(max_iterations):
					for str_filetype in cls.df_Original_headers.columns:
						df_myparamheaders = cls.df_Original_headers[cls.df_Original_headers[str_filetype] != '']
						lst_currentheaders = [str(col).upper().strip() for col in df_myparamheaders[str_filetype].tolist()]
						lst_mycurrentrow = [str(cell).upper().strip() for cell in df_mydf.iloc[int_myheaderrow].tolist()]
						if set(lst_currentheaders) == set(lst_mycurrentrow):
							lst_Findapnewheaders = [new_head for new_head in cls.df_Findap_newheaders[str_filetype].tolist() if new_head]
							lst_currentheaders_idx = [idx for idx, header in enumerate(lst_mycurrentrow) if header in lst_currentheaders]
							return str_filetype, int_myheaderrow, lst_currentheaders, lst_currentheaders_idx, lst_Findapnewheaders
				for int_myheaderrow in range(max_iterations):
					for str_filetype in cls.df_Original_headers.columns:
						df_myparamheaders = cls.df_Original_headers[cls.df_Original_headers[str_filetype] != '']
						lst_currentheaders = [str(col).upper().strip() for col in df_myparamheaders[str_filetype].tolist()]
						lst_mycurrentrow = [str(cell).upper().strip() for cell in df_mydf.iloc[int_myheaderrow].tolist()]
						
						if set(lst_currentheaders).issubset(set(lst_mycurrentrow)):
							lst_Findapnewheaders = [new_head for new_head in cls.df_Findap_newheaders[str_filetype].tolist() if new_head]
							lst_currentheaders_idx = [idx for idx, header in enumerate(lst_mycurrentrow) if header in lst_currentheaders]
							return str_filetype, int_myheaderrow, lst_currentheaders, lst_currentheaders_idx, lst_Findapnewheaders
				
				return 'UNKNOWN', None, [], [], []
			
			for sheet_name, df_mysheetdata in dic_myfile_dfs.items():
				dic_myfile_categories[sheet_name] = fn_find_matching_category(df_mysheetdata)
			
			return dic_myfile_categories
		
		except Exception as e:
			print(f'ERROR WHILE PROCESSING {obj_file} [{var_mysheet}].....ERROR = {e}')
			return {var_mysheet: ['UNKNOWN', None, [], [], []]}


	@staticmethod
	def fn_get_file_as_dicdataframes(str_filepath, var_mysheet=None, int_headerrow=None, nrows=35):
		"""Load an uploaded file (xls, xlsx, csv) into a DataFrame."""
		dic_mydfs={}
		
		if str_filepath.endswith('.csv'):
			dic_mydfs[0]=pd.read_csv(str_filepath, header=int_headerrow, nrows=nrows, na_filter=False)
		elif str_filepath.endswith('.xls'):
			dic_mydfs=pd.read_excel(str_filepath, sheet_name=var_mysheet, header=int_headerrow, nrows=nrows, na_filter=False, engine='xlrd')
			#dic_mydfs=pd.ExcelFile(str_filepath, engine='xlrd')
		elif str_filepath.endswith('.xlsx'):
			dic_mydfs=pd.read_excel(str_filepath, sheet_name=var_mysheet, header=int_headerrow, nrows=nrows, na_filter=False, engine='openpyxl')
		elif str_filepath.endswith('.xlsb'):
			dic_mydfs=pd.read_excel(str_filepath, sheet_name=var_mysheet, header=int_headerrow, nrows=nrows, na_filter=False, engine='pyxlsb')
		elif str_filepath.endswith('.xlsm'):
			dic_mydfs=pd.read_excel(str_filepath, sheet_name=var_mysheet, header=int_headerrow, nrows=nrows, na_filter=False, engine='openpyxl')
		
		return dic_mydfs
	
	@classmethod
	def load_date_formats(cls):
		"""Load date formats once for the entire class."""
		cls.load_parameters()
		return cls.lst_dates_regex
		
	@classmethod
	def fn_get_categories_sheetnames(cls):
		"""Get categories sheet names."""
		cls.load_parameters()
		return cls.df_Categoriesparams

	@classmethod
	def fn_get_categories_newheaders(cls):
		"""Get categories new headers."""
		cls.load_parameters()
		return cls.df_Findap_newheaders

	@classmethod
	def load_analysis_headers(cls, str_mysheetname):
		"""Load global analysis headers from AWS only once per execution."""
		cls.load_parameters()
		return cls.dic_filesparams_dfs[str_mysheetname]

	@classmethod
	def load_check_duplicate_params(cls):
		"""Load check duplicate parameters from AWS only once per execution."""
		cls.load_parameters()
		return cls.df_Check_duplicates_params

	@classmethod
	def load_finstatment_map(cls):
		"""Load FINDAP Trial Balance template parameters."""
		cls.load_parameters()
		return cls.df_finstatment_map

	@classmethod
	def load_TB_params(cls):
		"""Load FINDAP Trial Balance template parameters."""
		cls.load_parameters()
		return cls.df_TB_params

	@classmethod
	def fn_add_data2combined(cls, df_dataset, str_datasetcategory, str_mysheetname, str_headers_columnname):
		"""Transform dataset by applying column renaming and formula-based calculations."""
		cls.load_parameters()
		
		if df_dataset.empty:
			return pd.DataFrame()

		cls.load_analysis_headers(str_mysheetname)  # Load headers once

		# if str_datasetcategory not in cls.df_Findap_GlobalAnalysis_headers.columns:
		if str_datasetcategory not in cls.dic_filesparams_dfs[str_mysheetname].columns:
			return pd.DataFrame()  # Skip categories not in the analysis

		df_mydataset = df_dataset.fillna(0).replace(['NaN', ''], 0).copy()
		df_mydataset['CONCAT_SPACE'], df_mydataset['CONCAT_NORMAL_SPACE'], df_mydataset['CONCAT_SLASH'] = '_', ' ', '/'

		lst_mycategoryheaders = cls.dic_filesparams_dfs[str_mysheetname][str_datasetcategory].drop_duplicates().tolist()
		lst_Newheaders = cls.dic_filesparams_dfs[str_mysheetname][str_headers_columnname].drop_duplicates().tolist()

		lst_columns2keep = []
		for var_mycolumn in lst_mycategoryheaders:
			if '=' in var_mycolumn:
				str_mycolumnheader, str_mycolumnvalues = map(str.strip, var_mycolumn.split('=', 1))
				if '[' in str_mycolumnvalues:
					str_mycolumnvalues = str_mycolumnvalues.replace('[', '@df_mydataset["').replace(']', '"]').replace('""', '"')
			else:
				str_mycolumnheader, str_mycolumnvalues = var_mycolumn, ''

			if str_mycolumnheader and str_mycolumnheader not in df_mydataset.columns:
				if 'df_mydataset' in str_mycolumnvalues or 'astype' in str_mycolumnvalues:
					df_mydataset[str_mycolumnheader] = df_mydataset.eval(str_mycolumnvalues)
				elif 'float' in str_mycolumnvalues:
					df_mydataset[str_mycolumnheader] = float(str_mycolumnvalues.replace('float', '').replace('(', '').replace(')', ''))
				else:
					df_mydataset[str_mycolumnheader] = str_mycolumnvalues

			if str_mycolumnheader:
				lst_columns2keep.append(str_mycolumnheader)

		# Retain only the relevant columns
		df_mydataset = df_mydataset[lst_columns2keep]
		try:
			df_mydataset.columns = lst_Newheaders
		except Exception as e:
			print(f'ERROR WHILE NORMALIZING COLUMNS FOR DATASET {str_datasetcategory}.....ERROR = {e}')

		return df_mydataset

	@classmethod
	def fn_check_duplicatedrecords(cls, df_dataset, str_datasetcategory):
		"""Check for duplicate records based on predefined criteria."""
		if df_dataset.empty:
			return pd.DataFrame()

		cls.load_check_duplicate_params()  # Load parameters once

		df_mydataset = df_dataset.copy()
		lst_Criteriacolumns = cls.df_Check_duplicates_params.get(str_datasetcategory, pd.Series()).dropna().str.strip().tolist()

		# Use all columns if no specific criteria are found
		lst_Criteriacolumns = [col for col in lst_Criteriacolumns if col in df_mydataset.columns] or df_mydataset.columns.tolist()

		var_duplicate_mask = df_mydataset.duplicated(subset=lst_Criteriacolumns, keep='first')
		var_all_duplicates_mask = df_mydataset.duplicated(subset=lst_Criteriacolumns, keep=False)

		df_mydataset['Duplicate Status'] = 'NO duplicates'
		df_mydataset.loc[var_all_duplicates_mask & ~var_duplicate_mask, 'Duplicate Status'] = 'HAS duplicates'
		df_mydataset.loc[var_duplicate_mask, 'Duplicate Status'] = 'IS duplicate'

		return df_mydataset

	def fn_clean_illegal_characters(cell_value):
		# Clean illegal characters
		return re.sub(r'.', cls_Customfiles_Filetypehandler.fn_replace_ascii_illegalExcel, str(cell_value))

	def fn_replace_ascii_illegalExcel(match):
		char = match.group(0)
		if ord(char) == 8:
			return '<|>' 
		elif ord(char) == 160:
			return ''
		else:
			return char

	@classmethod
	def fn_convert_Uploadedfile2dataframe(cls,str_Folderpath, obj_file, str_Filecategory, str_Sheetname=0, int_headerrow=0,
										lst_currentheaders=[], lst_currentheadersindex=[], lst_Newheaders=[], dic_listofheaders2lookfor={}):
		# str_Paramfolder = 'FINDAP_FILES PARAMETERS/'
		# str_Paramfilename = 'FINDAP_Filetypes_Parameters.xlsx'
		df_Filedata = pd.DataFrame()
		str_Originfilename = str(obj_file).replace("<FileStorage: '", '').split("' ")[0]

		# Extend lst_currentheadersindex if needed
		lst_currentheadersindex.extend(range(len(lst_currentheadersindex)+1, len(lst_Newheaders)+1))

		# Determine Excel engine and load file
		if str(obj_file).endswith('.xls'):
			str_Excelreadengine = 'xlrd'
		elif str(obj_file).endswith('.xlsb'):
			str_Excelreadengine = 'pyxlsb'  
		else:
			str_Excelreadengine = 'openpyxl'  
			
		if str(obj_file)[-4:] == '.csv':
			df_mysheetdataAll = pd.read_csv(obj_file, header=int_headerrow, na_filter=False)
		else:
			str_temp_filepath = os.path.normpath(os.path.join(str_my_MEDIA_ROOT, str(obj_file)))
			if str(obj_file)[-4:] == '.xls':
				obj_file.seek(0)
				wbk_myworkbook=xlrd.open_workbook(file_contents=obj_file.read())
				wbk_tempfile = xlwt.Workbook()
				for int_sheet_index in range(wbk_myworkbook.nsheets):
					obj_mysheet = wbk_myworkbook.sheet_by_index(int_sheet_index)
					obj_newsheet = wbk_tempfile.add_sheet(obj_mysheet.name)
					for row in range(obj_mysheet.nrows):
						for col in range(obj_mysheet.ncols):
							obj_newsheet.write(row, col, obj_mysheet.cell_value(row, col))

				wbk_tempfile.save(str_temp_filepath)
				obj_tempfile= pd.ExcelFile(obj_file,engine=str_Excelreadengine)
			elif str(obj_file).endswith('.xlsb'):
				# wbk_tempfile = pyxlsb.open_workbook(obj_file,)
				obj_tempfile = pd.ExcelFile(str_temp_filepath, engine=str_Excelreadengine)
			else:
				wbk_tempfile = openpyxl.load_workbook(obj_file, data_only=True)
				wbk_tempfile.save(str_temp_filepath)
				obj_tempfile = pd.ExcelFile(str_temp_filepath, engine=str_Excelreadengine)
			df_mysheet = obj_tempfile.parse(str_Sheetname, header=None)

		if not lst_Newheaders: 
			lst_Newheaders = lst_currentheaders

		# Prepare headers to look for
		if not dic_listofheaders2lookfor:
			dic_listofheaders2lookfor[0] = lst_currentheaders
		lst_startrows_all = []

		for int_idx, headers in dic_listofheaders2lookfor.items():
			lst_currentheaders_upper = [str(h).upper() for h in headers]
			for idx, row in df_mysheet.iterrows():
				lst_mycurrentrow = [str(cell).upper() for cell in row[:250]]
				if set(lst_currentheaders_upper).issubset(lst_mycurrentrow):
					lst_startrows_all.append(idx)

		lst_startrows_all = sorted(set(lst_startrows_all))
		
		#if len(lst_startrows_all) == 1:
		if not (str_Filecategory in ['DHEgrp-Xl Sales V01','DHEgrp-Xl Sales V02','DHEgrp-Xl Sales V03']):
			df_mysheetdataAll = pd.read_excel(str_temp_filepath, sheet_name=str_Sheetname, header=None, 
											skiprows=int_headerrow + 1, usecols=lst_currentheadersindex, na_filter=False, engine=str_Excelreadengine)
		else:
			dic_DailySalestables = {}
			lst_startrows_all = lst_startrows_all[1:]
			df_mysheet = df_mysheet.iloc[:, lst_currentheadersindex]
			for i, int_dailystartrow in enumerate(lst_startrows_all):
				int_dailyendrow = lst_startrows_all[i + 1] - 4 if i < len(lst_startrows_all) - 1 else None
				if int_dailyendrow:  # If end row is defined
					df_mydailydata = df_mysheet.iloc[int_dailystartrow:int_dailyendrow, :]
				else:  # If end row is not defined, take all rows from start row to the end of the sheet
					df_mydailydata = df_mysheet.iloc[int_dailystartrow:, :]
				
				str_Transactiondate = df_mysheet.iloc[int_dailystartrow - 2, 0]
				try:
					dte_Transactiondate = datetime.strptime(str_Transactiondate.split(':')[1].strip(), '%d/%m/%Y').strftime('%Y-%m-%d')
				except:
					dte_Transactiondate = datetime(1900, 1, 1)
				
				int_lastcol = df_mydailydata.shape[1] - 1
				int_beforelastcol = df_mydailydata.shape[1] - 2

				# Assign the values to the respective columns
				df_mydailydata.iloc[:, int_beforelastcol] = str_Transactiondate
				df_mydailydata.iloc[:, int_lastcol] = dte_Transactiondate
				dic_DailySalestables[str_Transactiondate] = df_mydailydata

			df_mysheetdataAll = pd.concat(dic_DailySalestables.values(), keys=dic_DailySalestables.keys())

		if not lst_Newheaders: 
			lst_Newheaders = lst_currentheaders

		# var_spy=list(df_mysheetdataAll.columns)
		if not df_mysheetdataAll.empty:
			df_mysheetdataAll.columns = lst_Newheaders
			df_mysheetdataAll['ORIGIN_FILE'] = str_Originfilename
			df_mysheetdataAll['ORIGIN_SHEETNAME'] = str_Sheetname
			df_mysheetdataAll = df_mysheetdataAll.map(cls_Customfiles_Filetypehandler.fn_clean_illegal_characters)

		if not df_mysheetdataAll.empty:
			cls.load_parameters()
			df_mysheetdataAll['DATA GROUP'] = cls.df_Categoriesparams.loc[cls.df_Categoriesparams['CATEGORY'] == str_Filecategory, 'DATA GROUP'].values[0]
			df_mysheetdataAll['CATEGORY GROUP'] = cls.df_Categoriesparams.loc[cls.df_Categoriesparams['CATEGORY'] == str_Filecategory, 'CATEGORY GROUP'].values[0]
			df_mysheetdataAll['FINANCIAL STATEMENT GROUP'] = cls.df_Categoriesparams.loc[cls.df_Categoriesparams['CATEGORY'] == str_Filecategory, 'FINANCIAL STATEMENT GROUP'].values[0]

			# Format columns
			for col in cls.df_Fields_formats['FORMAT FLOAT'].tolist():
				if col in df_mysheetdataAll.columns:
					# df_mysheetdataAll[col] = pd.to_numeric(df_mysheetdataAll[col].astype(str).strip().replace({",": "", "'": ""}, regex=True), downcast='float', errors='coerce')
					df_mysheetdataAll[col] = pd.to_numeric(df_mysheetdataAll[col].astype(str).replace({",": "", "'": ""}, regex=True), downcast='float', errors='coerce')

			for col in cls.df_Fields_formats['FORMAT DATE'].tolist():
				if col in df_mysheetdataAll.columns:
					df_mysheetdataAll[col] = df_mysheetdataAll[col].apply(cls_Customfiles_Filetypehandler.fn_parse_dates_multipleformats)

			df_mydatafileanalysed = df_mysheetdataAll
			lst_headers_RRAdatascfile = ['SUPPLIERSTIN', 'CLIENTSTIN', 'SDCID', 'SDCRECEIPTTYPECOUNTER', 'INVOICENUMBER', 
										'TRANSACTION DATE', 'ITEMNAME', 'ITEMUNITPRICE', 'ITEMQUANTITY', 'ITEMTOTALPRICE', 
										'TAXTYPE', 'ORIGIN_FILE', 'ORIGIN_SHEETNAME']
			lst_Harmonizedheaders_RRAdatascfile = ['SUPPLIER TIN', 'CLIENT TIN', 'SDC ID', 'SDC COUNTER', 'SDC INVOICE NUMBER', 
												'TRANSACTION DATE', 'ITEM NAME', 'ITEM UNIT PRICE', 'ITEM QUANTITY', 
												'ITEM TOTAL PRICE', 'TAX TYPE', 'ORIGIN_FILE', 'ORIGIN_SHEETNAME']

			if str_Filecategory in ['RRAdsc PURCHASES-SALES V04']:
				df_mydatafileanalysed.drop(['INVOICETYPE'], axis=1, inplace=True)
			if str_Filecategory in ['RRAdsc PURCHASES-SALES V05']:
				df_mydatafileanalysed.drop(['SDCRECEIPTSIGNATURE'], axis=1, inplace=True)
			if str_Filecategory in ['RRAdsc PURCHASES-SALES V01', 'RRAdsc PURCHASES-SALES V03']:
				df_mydatafileanalysed['TAXTYPE'] = 'UNKNOWN'
			if str_Filecategory in ['RRAdsc PURCHASES-SALES V01']:
				df_mydatafileanalysed['SDCRECEIPTTYPECOUNTER'] = df_mydatafileanalysed['INVOICENUMBER'].str.split('/').str[1]
			if str_Filecategory in ['RRAdsc PURCHASES-SALES V01', 'RRAdsc PURCHASES-SALES V02', 'RRAdsc PURCHASES-SALES V03',
									'RRAdsc PURCHASES-SALES V04', 'RRAdsc PURCHASES-SALES V05']:
				df_mydatafileanalysed['INVOICENUMBER'] = df_mydatafileanalysed['SDCID'] + '/' + df_mydatafileanalysed['SDCRECEIPTTYPECOUNTER']
				df_mydatafileanalysed = df_mydatafileanalysed[lst_headers_RRAdatascfile]
				df_mydatafileanalysed.columns = lst_Harmonizedheaders_RRAdatascfile
		else:
			df_mydatafileanalysed = pd.DataFrame()

		return df_mydatafileanalysed


	@staticmethod
	def fn_convert_Ebmv21file2dataframe(obj_file,str_Filecategory):
		str_Paramfolder='FINDAP_FILES PARAMETERS/'
		str_Paramfilename='FINDAP_Filetypes_Parameters.xlsx'
		df_Filedata= pd.DataFrame()
		str_Originfilename =str(obj_file).replace("<FileStorage: '",'').replace("' ('application/vnd.ms-excel')>",'')
		str_Originfilename =str_Originfilename.replace("' ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')>",'')

		if str(obj_file)[-4:]=='.xls':
			df_myfile = pd.read_excel(obj_file,sheet_name=0, header=0,na_filter=False,engine='xlrd',)
		else:
			df_myfile = pd.read_excel(obj_file,sheet_name=0, header=0,na_filter=False,engine='openpyxl',)
			
		str_destination_filepath = os.path.normpath(os.path.join(str_my_MEDIA_ROOT, str_Paramfilename))
		if os.path.isfile(str_destination_filepath):
			os.remove(str_destination_filepath)
		cls_Aws_ParamfileHandler.get_fromaws_findap_filetypesparams(str_Paramfolder,str_Paramfilename,str_Paramfilename)
		df_Findapparamheaders = pd.read_excel(str_Paramfilename, sheet_name='FindapHeaders', na_filter=False,engine='openpyxl',)
			
		for str_filetype in df_Findapparamheaders.columns.tolist():
			if str_filetype==str_Filecategory:
				var_todrop=df_Findapparamheaders[df_Findapparamheaders[str_filetype]==''].index
				df_Findapparamheaders.drop(var_todrop , inplace=True)
				lst_Newheaders=df_Findapparamheaders[str_filetype].drop_duplicates().to_list()
				lst_Headers_Invoicedetails=df_Findapparamheaders[(str_filetype + '_DETAILS')].drop_duplicates().to_list()
				try:
					df_myfile.columns=lst_Newheaders
				except Exception as var_myError:
					print('ERROR WHILE RENAMING COLUMNS FOR {}.....ERROR = {}'.format(str(obj_file),var_myError))
					return df_myfile
		str_Cdtion_Invoicesummary=(df_myfile['REGDT']!='') & (df_myfile['REGDT']!='RegDt') & (df_myfile['ORGINVCNO']!='OrgInvcNo')
		str_Cdtion_Invoicedetails=(df_myfile['ORGINVCNO']!='') & (df_myfile['ORGINVCNO']!='ItemSeq') & (df_myfile['ORGINVCNO']!='0') & (df_myfile['ORGINVCNO']!=0) & (df_myfile['ORGINVCNO']!='ItemCd') & (df_myfile['ORGINVCNO']!='OrgInvcNo')
		df_Invoicesummary = df_myfile.loc[str_Cdtion_Invoicesummary] 
		df_Invoicedetails = df_myfile.loc[str_Cdtion_Invoicedetails] 
		df_Invoicedetails.columns=lst_Headers_Invoicedetails
		
		if len(df_Invoicesummary)!=0:
			str_destination_filepath = os.path.normpath(os.path.join(str_my_MEDIA_ROOT, str_Paramfilename))
			if os.path.isfile(str_destination_filepath):
				os.remove(str_destination_filepath)
			cls_Aws_ParamfileHandler.get_fromaws_findap_filetypesparams(str_Paramfolder,str_Paramfilename,str_Paramfilename)
			df_Dataformat = pd.read_excel(str_Paramfilename, sheet_name='Dataformat', na_filter=False,engine='openpyxl',)

			#sFormatting FLOAT
			lst_Col2format_Summary= [value for value in df_Invoicesummary.columns.to_list() if value in df_Dataformat['FORMAT FLOAT'].to_list()]
			for col in lst_Col2format_Summary:
				try:
					df_Invoicesummary[col]= df_Invoicesummary[col].replace({",": ""},regex=True)
					df_Invoicesummary[col]= pd.to_numeric(df_Invoicesummary[col], downcast='float')
				except Exception as var_myError:
					print('UNABLE TO FORMAT COLUMN {} INTO FLOAT.....ERROR = {}'.format(str(col),var_myError))
					pass

			lst_Col2format_Details= [value for value in df_Invoicedetails.columns.to_list() if value in df_Dataformat['FORMAT FLOAT'].to_list()]
			for col in lst_Col2format_Details:
				try:
					df_Invoicedetails[col]= df_Invoicedetails[col].replace({",": ""},regex=True)
					df_Invoicedetails[col]= pd.to_numeric(df_Invoicedetails[col], downcast='float')
				except Exception as var_myError:
					print('UNABLE TO FORMAT COLUMN {} INTO FLOAT.....ERROR = {}'.format(str(col),var_myError))
					pass

			#Formatting DATE
			lst_Col2format_Summary= [value for value in df_Invoicesummary.columns.to_list() if value in df_Dataformat['FORMAT DATE'].to_list()]
			for col in lst_Col2format_Summary:
				try:
					#df_Invoicesummary[col]= pd.to_datetime(df_Invoicesummary[col],errors='coerce', infer_datetime_format=True,)
					df_Invoicesummary[col]= df_Invoicesummary[col].apply(cls_Customfiles_Filetypehandler.fn_parse_dates_multipleformats)
				except Exception as var_myError:
					print('UNABLE TO FORMAT COLUMN {} INTO DATE.....ERROR = {}'.format(str(col),var_myError))
					pass

			lst_Col2format_Details= [value for value in df_Invoicedetails.columns.to_list() if value in df_Dataformat['FORMAT DATE'].to_list()]
			for col in lst_Col2format_Details:
				try:
					#df_Invoicedetails[col]= pd.to_datetime(df_Invoicedetails[col],errors='coerce', infer_datetime_format=True,)
					df_Invoicedetails[col]= df_Invoicedetails[col].apply(cls_Customfiles_Filetypehandler.fn_parse_dates_multipleformats)
				except Exception as var_myError:
					print('UNABLE TO FORMAT COLUMN {} INTO DATE.....ERROR = {}'.format(str(col),var_myError))
					pass
			df_myEbmv21file_Summary = df_Invoicesummary
			df_myEbmv21file_Details = df_Invoicedetails
		else:
			df_myEbmv21file_Summary = pd.DataFrame()
			df_myEbmv21file_Details = pd.DataFrame()
		
		df_myEbmv21file_Summary=df_myEbmv21file_Summary.assign(ORIGIN_FILE=str_Originfilename)
		df_myEbmv21file_Details=df_myEbmv21file_Details.assign(ORIGIN_FILE=str_Originfilename)
		return [df_myEbmv21file_Summary,df_myEbmv21file_Details]

	@staticmethod
	def fn_split_Ebmreceipt2sdcnum_invoicenum(str_ReceiptNumber):
		str_Sdc=str_ReceiptNumber.split('/')[0]
		str_Invoicenumber=str_ReceiptNumber.split('/')[1]
		lng_Invoicenumber=int(str_Invoicenumber.replace('.',''))
		return str_Sdc,lng_Invoicenumber,None

	@staticmethod
	def fn_get_Keywords_representativity(df_mydataframe,str_Columndescription,str_Columnamount):
		print('GETTING KEYWORDS FRAQUENCIES & RELATED AMOUNT.....')
		dic_Keywords = {}
		df_Keywords_representativity = pd.DataFrame()
		for index, row in df_mydataframe.iterrows():
			str_words = row[str_Columndescription].split()
			try:
				flt_Amount = float(row[str_Columnamount])
			except:
				flt_Amount=0
				pass
			for str_word in str_words:
				str_word = str_word.lower()
				try:
					int_wordcount = dic_Keywords[str_word][0] + 1
					flt_Cumulamount = dic_Keywords[str_word][1] + flt_Amount
				except:
					int_wordcount = 1
					flt_Cumulamount =0

				dic_Keywords[str_word] = [int_wordcount,flt_Cumulamount]
		df_Keywords_representativity = pd.DataFrame(dic_Keywords).T
		df_Keywords_representativity.columns = ['FREQUENCY','RELATED AMOUNT']
		return df_Keywords_representativity

	def safe_eval(col_value, df):
		"""
		Safely evaluates specific expressions in `col_value`.
		Handles type casting, DataFrame operations, and literal evaluations.
		"""
		try:
			if 'astype' in col_value:
				# Extract type and apply astype conversion
				target_type = col_value.split('astype(')[1].strip(')').strip()
				return df.astype(target_type)
			elif col_value.strip().replace('.', '', 1).isdigit():
				# Direct numeric conversion
				return float(col_value) if '.' in col_value else int(col_value)
			elif col_value.startswith("'") or col_value.startswith('"'):
				# Return string literal
				return col_value.strip("'\"")
			elif 'df_mydataset' in col_value:
				# Evaluate expression involving the DataFrame
				return df.eval(col_value.strip())
			else:
				# Safely evaluate basic Python literals
				return ast.literal_eval(col_value.strip())
		except Exception as e:
			print(f"Error evaluating expression '{col_value}': {e}")
			return ''


	def fn_extract_Ebmreferences_Sales(str_mynote):
		#match = re.search(r'invoice\s+(\d+)', str_mynote, re.IGNORECASE)
		match = re.search(r'Invoice[:\str_mynote]*(.*)', str_mynote, re.IGNORECASE)
		if match:
			str_after_invoice = match.group(1)
			# Use regex to find all numeric substrings, including those with periods and commas
			return re.findall(r'\d+(?:[.,/]\d+)?', str_after_invoice)
		else:
			return 'Missing'
	
	def fn_extract_Ebmreferences(str_mynote,str_Category,str_mykeyword):
		if str_Category in ['SALES']:
			#match = re.search(r'Invoice[:\str_mynote]*(.*)', str_mynote, re.IGNORECASE)
			var_match = re.search(rf'{str_mykeyword}[:\str_mynote]*(.*)', str_mynote, re.IGNORECASE)
			if var_match:
				str_after_invoice = var_match.group(1)
				# Use regex to find all numeric substrings, including those with periods and commas
				return re.findall(r'\d+(?:[.,/]\d+)?', str_after_invoice)
			else:
				return 'Missing'
			
		elif str_Category in ['PURCHASES']:
			var_matches = re.findall(rf'{str_mykeyword}\s*([\d\s&/\.,]+)\s*', str_mynote, re.IGNORECASE)
			if var_matches:
				return '; '.join([f'SDC{str_num}' for str_num in var_matches])
			else:
				return 'Missing'

	def fn_insert_missingdates(df_dataset, str_datecolumn, lst_othercolumns=None):
		if lst_othercolumns is None:
			lst_othercolumns = []

		# Ensure the combination of date column and other columns is unique
		df_dataset.drop_duplicates(subset=[str_datecolumn] + lst_othercolumns, inplace=True)
		df_dataset[str_datecolumn] = pd.to_datetime(df_dataset[str_datecolumn])

		# Set the date column as the index
		df_dataset.set_index([str_datecolumn] + lst_othercolumns, inplace=True)

		# Create a full date range
		var_datesfullrange = pd.date_range(start=df_dataset.index.get_level_values(str_datecolumn).min(),
										end=df_dataset.index.get_level_values(str_datecolumn).max(), freq='D')
		var_datesfullrange = pd.to_datetime(var_datesfullrange)

		# Reindex to fill missing dates for all combinations of other columns
		df_dataset = df_dataset.reindex(pd.MultiIndex.from_product([var_datesfullrange] + 
																[df_dataset.index.get_level_values(col).unique() for col in lst_othercolumns],
																names=[str_datecolumn] + lst_othercolumns))

		# Forward fill missing data
		df_dataset.ffill(inplace=True)
		df_dataset = df_dataset.infer_objects(copy=False)	# Infer object types after fill
		pd.set_option('future.no_silent_downcasting', True)

		# Reset the index
		df_dataset.reset_index(inplace=True)

		return df_dataset
	

	def fn_exchange_negativesign(df_dataset,str_fromcolumn,str_tocolumn):
		def fn_process_exchange(row):
			if row[str_fromcolumn] < 0:
				row[str_fromcolumn], row[str_tocolumn] = abs(row[str_fromcolumn]), -abs(row[str_tocolumn])
			return row

		df_dataset[[str_fromcolumn, str_tocolumn]] = df_dataset.apply(fn_process_exchange, axis=1)[[str_fromcolumn, str_tocolumn]]
		return df_dataset

	def fn_get_invoicesdfs(df_dataset,str_criteriafield,lst_Tin2remove,lst_Header_invoicesummary,lst_Header_invoicedetails,lst_col2keep,dic_key_fields2add,dic_fields2add):
		# Assuming df_dataset is your dataframe and lst_Tin2remove is your list
		df_dataset = df_dataset[~df_dataset[str_criteriafield].isin(lst_Tin2remove)]
		df_dataset = df_dataset[df_dataset[str_criteriafield].notna()]
		df_dataset = df_dataset[df_dataset[str_criteriafield] != '']

		# Define a function to check if a variable date is valid
		def fn_is_valid_date(var_date):
			if var_date in [None, '', 'NaT', 'Nan','NaN',]:
				return False
			try:
				dte_mydate = pd.to_datetime(var_date)
				if dte_mydate != datetime(1900, 1, 1):
					return True
				else:
					return False
			except (ValueError, TypeError):
				return False
				
		# Apply the function to create a boolean mask to filter invoices summary
		var_valid_date_mask = df_dataset['TRANSACTION DATE'].apply(fn_is_valid_date)
		df_Invoicesummary = df_dataset[var_valid_date_mask].copy()
		df_Invoicesummary.columns = lst_Header_invoicesummary

		# Create boolean masks for the conditions
		var_notempty_mask = df_dataset[str_criteriafield].notna() & (df_dataset[str_criteriafield] != '')
		var_combined_mask = var_notempty_mask & (~var_valid_date_mask)
		df_Invoicedetails = df_dataset[var_combined_mask].copy()
				
		# Check the number of current columns and the new columns
		int_num_current_cols = len(df_Invoicedetails.columns)
		int_num_new_cols = len(lst_Header_invoicedetails)

		if int_num_current_cols > int_num_new_cols:
			df_Invoicedetails = df_Invoicedetails.iloc[:, :int_num_new_cols]
		elif int_num_current_cols < int_num_new_cols:
			# Add missing columns and fill with default value '-'
			for int_count in range(int_num_new_cols - int_num_current_cols):
				new_col_name = lst_Header_invoicedetails[int_num_current_cols + int_count]
				df_Invoicedetails[new_col_name] = '-'

		# Rename columns to new headers 
		df_Invoicedetails.columns = lst_Header_invoicedetails

		for str_col in lst_col2keep:
			if str_col not in dic_fields2add:
				dic_fields2add[str_col] = str_col

		df_Invoicedetails = cls_Customfiles_Filetypehandler.fn_update_invoicedetailsfromsummary(df_Invoicesummary, df_Invoicedetails, dic_key_fields2add, dic_fields2add)

		df_Invoicesummary = cls_Customfiles_Filetypehandler.fn_apply_formats(df_Invoicesummary)
		df_Invoicedetails = cls_Customfiles_Filetypehandler.fn_apply_formats(df_Invoicedetails)
		return df_Invoicesummary,df_Invoicedetails

	def fn_update_invoicedetailsfromsummary(df_Invoicesummary, df_Invoicedetails, dic_key_fields2add, dic_fields2add):
		# Check for the existence of key columns
		bln_key_fields_exist = all([str_key in df_Invoicesummary.columns and str_value in df_Invoicedetails.columns for str_key, str_value in dic_key_fields2add.items()])

		if bln_key_fields_exist:
			# Add missing columns to df_Invoicedetails
			for str_summary_col, str_details_col in dic_fields2add.items():
				if str_details_col not in df_Invoicedetails.columns:
					df_Invoicedetails[str_details_col] = '-'
			
			# Perform the update based on key fields
			for var_Idx, var_Row in df_Invoicedetails.iterrows():
				conditions = [df_Invoicesummary[summary_key] == var_Row[detail_key] for summary_key, detail_key in dic_key_fields2add.items()]
				var_matching_rows = df_Invoicesummary[conditions[0]]
				for condition in conditions[1:]:
					var_matching_rows = var_matching_rows[condition]
				
				if not var_matching_rows.empty:
					for str_summary_col, str_details_col in dic_fields2add.items():
						df_Invoicedetails.at[var_Idx, str_details_col] = var_matching_rows.iloc[0][str_summary_col]
		else:
			print("Required key columns are missing in one of the dataframes. No update performed.")
		
		return df_Invoicedetails

	@classmethod
	def fn_apply_formats(cls,df_dataset):
		cls.load_parameters()
		# Convert lists of format columns
		format_float_columns = set(cls.df_Fields_formats['FORMAT FLOAT'].tolist())
		format_date_columns = set(cls.df_Fields_formats['FORMAT DATE'].tolist())

		# Determine which columns need formatting
		lst_float_columns_to_format = list(format_float_columns.intersection(df_dataset.columns))
		lst_date_columns_to_format = list(format_date_columns.intersection(df_dataset.columns))

		# Format columns
		if lst_float_columns_to_format:
			df_dataset[lst_float_columns_to_format] = df_dataset[lst_float_columns_to_format].replace({",": "", "'": ""}, regex=True).apply(pd.to_numeric, downcast='float', errors='coerce')

		if lst_date_columns_to_format:
			df_dataset[lst_date_columns_to_format] = df_dataset[lst_date_columns_to_format].map(cls_Customfiles_Filetypehandler.fn_parse_dates_multipleformats)

		return df_dataset

	def fn_calculate_tradescategory(df_dataset,):
		str_Cdtion_01=(df_dataset['TRANSACTION DATE'].notna())
		str_Cdtion_02=(df_dataset['TRANSACTION DATE']!='')
		str_Cdtion_All=str_Cdtion_01 & str_Cdtion_02 
		df_dataset = df_dataset.loc[str_Cdtion_All]
		if ('SUPPLIER TIN' in df_dataset.columns) and ('CLIENT TIN' in df_dataset.columns):
			if df_dataset['SUPPLIER TIN'].nunique()==1:
				df_dataset['TRANSACTION CATEGORY'] = 'SALES'
				df_dataset['FINANCIAL STATEMENT GROUP'] = 'SALES'
			elif df_dataset['CLIENT TIN'].nunique()==1:
				df_dataset['TRANSACTION CATEGORY'] = 'PURCHASES'
				df_dataset['FINANCIAL STATEMENT GROUP'] = 'COGS-EXPENSES'
		else:
			df_dataset['TRANSACTION CATEGORY'] = 'UNKNOWN'
			df_dataset['FINANCIAL STATEMENT GROUP'] = 'UNKNOWN'
		return df_dataset

	def fn_calculate_refund_status(df_dataset):
		# Step 1: Assign "REFUND" status for rows where OPERATION TYPE is "NR"
		df_dataset['REFUND STATUS'] = None
		df_dataset.loc[df_dataset['OPERATION TYPE'] == 'NR', 'REFUND STATUS'] = 'REFUND'

		# Step 2: Calculate group totals for 'MRC NUMBER' and add to the dataset
		group_totals = df_dataset.groupby('MRC NUMBER')[['VAT AMOUNT', 'AMOUNT VAT INCLUDED']].sum().reset_index()
		group_totals.rename(columns={
			'VAT AMOUNT': 'VAT AMOUNT_by_MRC NUMBER',
			'AMOUNT VAT INCLUDED': 'AMOUNT VAT INCLUDED_by_MRC NUMBER'
		}, inplace=True)
		
		df_dataset = df_dataset.merge(group_totals, on='MRC NUMBER', how='left')

		# Step 3: Update lst_matching_columns to include the new columns
		lst_matching_columns = ['ITEM CODE', 'ITEM NAME', 'QUANTITY', 'UNIT PRICE', 'BUYER TIN','SDC ID', 'VAT AMOUNT',
						'AMOUNT VAT INCLUDED', 'VAT AMOUNT_by_MRC NUMBER', 'AMOUNT VAT INCLUDED_by_MRC NUMBER'
						]

		# Step 4: Extract rows for "NR" and "NS"
		df_rows_NR = df_dataset[df_dataset['OPERATION TYPE'] == 'NR'].copy()
		df_rows_NS = df_dataset[df_dataset['OPERATION TYPE'] == 'NS'].copy()

		# Step 5: Add temporary column "Index_NR" for unique indexing in df_rows_NR
		df_rows_NR['Index_NR'] = df_rows_NR.index

		# Step 6: Merge "NR" and "NS" rows on matching columns
		df_merged = df_rows_NR.merge(df_rows_NS, on=lst_matching_columns, suffixes=('_NR', '_NS'), how='inner')

		# Retain Index_NR in the merged DataFrame
		df_merged['Index_NR'] = df_merged['Index_NR']

		# Step 7: Add Rank_NR to count duplicates of rows from df_rows_NR
		df_merged['Rank_NR'] = df_merged.groupby(lst_matching_columns).cumcount() + 1

		# Step 8: Keep only rows with Rank_NR = 1
		df_merged = df_merged[df_merged['Rank_NR'] == 1]

		# Step 9: Assign "REFUNDED" for rows in valid matches
		ns_matched_indices = df_merged['RECEIPT NUMBER (SDC)_NS']
		df_dataset.loc[df_dataset['RECEIPT NUMBER (SDC)'].isin(ns_matched_indices), 'REFUND STATUS'] = 'REFUNDED'

		# Step 10: Assign "NORMAL SALE" for remaining "NS" rows
		df_dataset.loc[(df_dataset['OPERATION TYPE'] == 'NS') & (df_dataset['REFUND STATUS'].isnull()), 'REFUND STATUS'] = 'NORMAL SALE'

		# Step 11: Assign "REFUND NO MATCH" for "NR" rows that do not have a matching "NS" row
		unmatched_NR_indices = df_rows_NR[~df_rows_NR['Index_NR'].isin(df_merged['Index_NR'])]
		df_dataset.loc[unmatched_NR_indices['Index_NR'], 'REFUND STATUS'] = 'REFUND NO MATCH'

		# Clean up temporary columns
		df_dataset.drop(columns=['Index_NR'], inplace=True, errors='ignore')

		return df_dataset

	def fn_update_transactiondate_ifNA (df_dataset):
		if not df_dataset.empty:
			# Function to extract and format the end date
			def fn_extract_and_format_enddate(str_originfile):
				var_match = re.search(r'ETAX_PAYE.*_(\d{9})_[^_]*_(\d{8})_(\d{8})_(\d{5,12}).*', str_originfile)
				if var_match:
					dte_enddate = var_match.group(3)
					dte_formatted_enddate = datetime.strptime(dte_enddate, "%Y%m%d").strftime("%Y-%m-%d")
					return dte_formatted_enddate
				return None
			# Update TRANSACTION_DATE if the value is NA or not a date
			def fn_update_transactiondate(var_transactiondate, dte_formatted_enddate):
				try:
					# Try to parse the transaction_date to check if it's valid
					if (isinstance(var_transactiondate, str) and var_transactiondate.isdigit()) :
						var_transactiondate = datetime(1899, 12, 30) + timedelta(days=int(var_transactiondate))
					pd.to_datetime(var_transactiondate)
				except (ValueError, TypeError):
					# If it's not a valid date, update it with the formatted end date
					return dte_formatted_enddate
				return var_transactiondate

			df_dataset['FORMATTED_END_DATE'] = df_dataset['ORIGIN_FILE'].apply(fn_extract_and_format_enddate)
			df_dataset['TRANSACTION DATE'] = df_dataset.apply(lambda row: fn_update_transactiondate(row['TRANSACTION DATE'], row['FORMATTED_END_DATE']), axis=1)
			df_dataset = df_dataset.drop(columns=['FORMATTED_END_DATE'])
		return df_dataset

	# Define the mapping dictionary for multiple languages
	dic_month_map = {
		'Jan': '01', 'January': '01', 'JANUARY': '01', 'JAN': '01','MUTARAMA': '01',
		'Feb': '02', 'Februar': '02', 'February': '02', 'Février': '02', 'Fév': '02', 'Fev': '02', 'Febrero': '02', 'FEB': '02', 'FEBRUARY': '02', 'FÉVRIER': '02', 'FÉV': '02', 'FEV': '02', 'FEBRERO': '02', 'FEBRUAR': '02','GASHYANTARE': '02',
		'Mar': '03', 'March': '03', 'März': '03', 'Mars': '03', 'Marzo': '03', 'MAR': '03', 'MARCH': '03', 'MÄRZ': '03', 'MARS': '03', 'MARZO': '03','WERURWE': '03',
		'Apr': '04', 'April': '04', 'Avril': '04', 'Abril': '04', 'APR': '04', 'APRIL': '04', 'AVRIL': '04', 'ABRIL': '04', 'MATA': '04',
		'May': '05', 'Mai': '05', 'Mayo': '05', 'MAY': '05', 'MAI': '05', 'MAYO': '05','GICURASI': '05',
		'Jun': '06', 'June': '06', 'Juni': '06', 'Juin': '06', 'Junio': '06', 'JUN': '06', 'JUNE': '06', 'JUNI': '06', 'JUIN': '06', 'JUNIO': '06','KAMENA': '06',
		'Jul': '07', 'July': '07', 'Juli': '07', 'Juillet': '07', 'Julio': '07', 'JUL': '07', 'JULY': '07', 'JULI': '07', 'JUILLET': '07', 'JULIO': '07','NYAKANGA': '07',
		'Aug': '08', 'August': '08', 'Août': '08', 'Agosto': '08', 'Aou': '08', 'AUG': '08', 'AUGUST': '08', 'AOÛT': '08', 'AGOSTO': '08','KANAMA': '08',
		'Sep': '09', 'September': '09', 'Septembre': '09', 'Sept': '09', 'Setiembre': '09', 'Septiembre': '09', 'SEP': '09', 'SEPTEMBER': '09', 'SEPTEMBRE': '09', 'SEPT': '09', 'SETIEMBRE': '09', 'SEPTIEMBRE': '09','NZERI': '09','NZELI': '09',
		'Oct': '10', 'October': '10', 'Oktober': '10', 'Octobre': '10', 'Octubre': '10', 'OCT': '10', 'OCTOBER': '10', 'OKTOBER': '10', 'OCTOBRE': '10', 'OCTUBRE': '10','UKWAKIRA': '10',
		'Nov': '11', 'November': '11', 'Novembre': '11', 'Noviembre': '11', 'NOV': '11', 'NOVEMBER': '11', 'NOVEMBRE': '11', 'NOVIEMBRE': '11','UGUSHYINGO': '11',
		'Dec': '12', 'December': '12', 'Dezember': '12', 'Décembre': '12', 'Diciembre': '12', 'Déc': '12', 'Dez': '12', 'DIC': '12', 'DEC': '12', 'DECEMBER': '12', 'DEZEMBER': '12', 'DÉCEMBRE': '12', 'DÉC': '12', 'DEZ': '12', 'DICIEMBRE': '12', 'UKUBOZA': '12',
	}
	# Compile the regex pattern for month replacements
	month_pattern = re.compile(r'\b(' + '|'.join(dic_month_map.keys()) + r')\b', re.IGNORECASE)
	
	# Function to replace month names and abbreviations
	@classmethod
	def fn_replace_months(cls_myclass, dte_mydate):
		return cls_myclass.month_pattern.sub(lambda x: cls_myclass.dic_month_map[x.group().upper()], dte_mydate)

	# Function to parse dates in multiple formats
	@classmethod
	def fn_parse_dates_multipleformats(cls_myclass, dte_mydate):
		if isinstance(dte_mydate, datetime):
			return dte_mydate

		if pd.isna(dte_mydate) or pd.isnull(dte_mydate) or dte_mydate in [None, '', 'NaN', 'NaT']:
			return dte_mydate  # dte_defaultdate

		# Handle numeric date formats from Excel
		# if isinstance(dte_mydate, (int, float)) or (isinstance(dte_mydate, str) and dte_mydate.isdigit()):
		if isinstance(dte_mydate, (int, float)) or (isinstance(dte_mydate, str) and dte_mydate.isdigit()):
			dte_mydate = str(int(float(dte_mydate)))  # Convert to integer string
			if int(dte_mydate) < 2958466:  # Check valid Excel date range
				dte_mydate = datetime(1899, 12, 30) + timedelta(days=int(dte_mydate))
				return dte_mydate

		# Replace month names with numeric values
		dte_mydate = cls_myclass.fn_replace_months(str(dte_mydate))

		# Ensure date formats are loaded
		cls_myclass.load_date_formats()

		# Try to parse the date using known formats
		for str_myformat in cls_myclass.lst_dates_regex:
			try:
				if isinstance(dte_mydate, str):
					return datetime.strptime(dte_mydate.strip(), str_myformat.strip())
				# print('My date {} format is {} and CONVERTED DATE is : {}'.format(dte_mydate,str_myformat.strip(),dte_mydate_formated))
				return datetime.strptime(dte_mydate, str_myformat.strip())
			except (ValueError, TypeError):
				continue

		# If no valid date format is found, return the original input
		return dte_mydate  # dte_defaultdate
	
	@classmethod
	def fn_convert_Worksheet2dataframe(cls,str_Folderpath, obj_file, str_Filecategory, str_Sheetname=0, int_headerrow=0,
										lst_currentheaders=[], lst_currentheadersindex=[], lst_Newheaders=[], dic_listofheaders2lookfor={}):
		cls.load_parameters()  # Load shared parameters if not already loaded
		str_Originfilename = str(obj_file).replace("<FileStorage: '", '').split("' ")[0]

		# Extend lst_currentheadersindex if needed
		lst_currentheadersindex.extend(range(len(lst_currentheadersindex) + 1, len(lst_Newheaders) + 1))
		if str(obj_file).endswith('.xls'):
			str_Excelreadengine = 'xlrd'  
		elif str(obj_file).endswith('.xlsb'):
			str_Excelreadengine = 'pyxlsb'  
		else:
			str_Excelreadengine ='openpyxl'

		if str(obj_file)[-4:] == '.csv':
			df_mysheetdataAll = pd.read_csv(obj_file, header=int_headerrow, na_filter=False)
		else:
			obj_tempfile = pd.ExcelFile(obj_file, engine=str_Excelreadengine)
			df_mysheet = obj_tempfile.parse(str_Sheetname, header=None)

		if not lst_Newheaders:
			lst_Newheaders = lst_currentheaders

		# Prepare headers to look for
		if not dic_listofheaders2lookfor:
			dic_listofheaders2lookfor[0] = lst_currentheaders

		lst_startrows_all = []
		for int_idx, headers in dic_listofheaders2lookfor.items():
			lst_currentheaders_upper = [str(h).upper() for h in headers]
			for idx, row in df_mysheet.iterrows():
				lst_mycurrentrow = [str(cell).upper() for cell in row[:250]]
				if set(lst_currentheaders_upper).issubset(lst_mycurrentrow):
					lst_startrows_all.append(idx)

		lst_startrows_all = sorted(set(lst_startrows_all))

		# Read and clean data
		if str_Filecategory not in ['DHEgrp-Xl Sales V01', 'DHEgrp-Xl Sales V02', 'DHEgrp-Xl Sales V03']:
			df_mysheetdataAll = pd.read_excel(obj_file, sheet_name=str_Sheetname, header=None,
												skiprows=int_headerrow + 1, usecols=lst_currentheadersindex, na_filter=False, engine=str_Excelreadengine)
		else:
			df_mysheetdataAll = pd.DataFrame()
			for i, int_dailystartrow in enumerate(lst_startrows_all[1:]):
				int_dailyendrow = lst_startrows_all[i + 1] - 4 if i < len(lst_startrows_all) - 1 else None
				df_mydailydata = df_mysheet.iloc[int_dailystartrow:int_dailyendrow, lst_currentheadersindex] if int_dailyendrow else df_mysheet.iloc[int_dailystartrow:, lst_currentheadersindex]

				str_Transactiondate = df_mysheet.iloc[int_dailystartrow - 2, 0]
				try:
					dte_Transactiondate = datetime.strptime(str_Transactiondate.split(':')[1].strip(), '%d/%m/%Y').strftime('%Y-%m-%d')
				except:
					dte_Transactiondate = '1900-01-01'

				if not df_mydailydata.empty:
					df_mydailydata.loc[:, 'TRANSACTION DATE'] = dte_Transactiondate
				df_mysheetdataAll = pd.concat([df_mysheetdataAll, df_mydailydata])

		# Standardize headers and clean data
		if not df_mysheetdataAll.empty:
			df_mysheetdataAll.columns = lst_Newheaders
			if 'TRANSACTION DATE' not in df_mysheetdataAll.columns:
				df_mysheetdataAll['TRANSACTION DATE'] = datetime(1900, 1, 1)
			df_mysheetdataAll['TRANSACTION DATE'] = df_mysheetdataAll['TRANSACTION DATE'].map(cls_Customfiles_Filetypehandler.fn_parse_dates_multipleformats)
			df_mysheetdataAll['ORIGIN_FILE'] = str_Originfilename
			df_mysheetdataAll['ORIGIN_SHEETNAME'] = str_Sheetname
			df_mysheetdataAll = df_mysheetdataAll.map(cls_Customfiles_Filetypehandler.fn_clean_illegal_characters)

		# Apply formatting
		if not df_mysheetdataAll.empty:
			df_mysheetdataAll['DATA GROUP'] = cls.df_Categoriesparams.loc[cls.df_Categoriesparams['CATEGORY'] == str_Filecategory, 'DATA GROUP'].values[0]
			df_mysheetdataAll['CATEGORY GROUP'] = cls.df_Categoriesparams.loc[cls.df_Categoriesparams['CATEGORY'] == str_Filecategory, 'CATEGORY GROUP'].values[0]
			df_mysheetdataAll['FINANCIAL STATEMENT GROUP'] = cls.df_Categoriesparams.loc[cls.df_Categoriesparams['CATEGORY'] == str_Filecategory, 'FINANCIAL STATEMENT GROUP'].values[0]

			# Format columns
			for col in cls.df_Fields_formats['FORMAT FLOAT'].tolist():
				if col in df_mysheetdataAll.columns:
					# df_mysheetdataAll[col] = pd.to_numeric(df_mysheetdataAll[col].astype(str).strip().replace({",": "", "'": ""}, regex=True), downcast='float', errors='coerce')
					df_mysheetdataAll[col] = pd.to_numeric(df_mysheetdataAll[col].astype(str).replace({",": "", "'": ""}, regex=True), downcast='float', errors='coerce')


			for col in cls.df_Fields_formats['FORMAT DATE'].tolist():
				if col in df_mysheetdataAll.columns:
					df_mysheetdataAll[col] = df_mysheetdataAll[col].apply(cls_Customfiles_Filetypehandler.fn_parse_dates_multipleformats)

			df_mydatafileanalysed = df_mysheetdataAll
			lst_headers_RRAdatascfile = ['SUPPLIERSTIN', 'CLIENTSTIN', 'SDCID', 'SDCRECEIPTTYPECOUNTER', 'INVOICENUMBER', 
										'TRANSACTION DATE', 'ITEMNAME', 'ITEMUNITPRICE', 'ITEMQUANTITY', 'ITEMTOTALPRICE', 
										'TAXTYPE', 'ORIGIN_FILE', 'ORIGIN_SHEETNAME']
			lst_Harmonizedheaders_RRAdatascfile = ['SUPPLIER TIN', 'CLIENT TIN', 'SDC ID', 'SDC COUNTER', 'SDC INVOICE NUMBER', 
												'TRANSACTION DATE', 'ITEM NAME', 'ITEM UNIT PRICE', 'ITEM QUANTITY', 
												'ITEM TOTAL PRICE', 'TAX TYPE', 'ORIGIN_FILE', 'ORIGIN_SHEETNAME']

			if str_Filecategory in ['RRAdsc PURCHASES-SALES V04']:
				df_mydatafileanalysed.drop(['INVOICETYPE'], axis=1, inplace=True)
			if str_Filecategory in ['RRAdsc PURCHASES-SALES V05']:
				df_mydatafileanalysed.drop(['SDCRECEIPTSIGNATURE'], axis=1, inplace=True)
			if str_Filecategory in ['RRAdsc PURCHASES-SALES V01', 'RRAdsc PURCHASES-SALES V03']:
				df_mydatafileanalysed['TAXTYPE'] = 'UNKNOWN'
			if str_Filecategory in ['RRAdsc PURCHASES-SALES V01']:
				df_mydatafileanalysed['SDCRECEIPTTYPECOUNTER'] = df_mydatafileanalysed['INVOICENUMBER'].str.split('/').str[1]
			if str_Filecategory in ['RRAdsc PURCHASES-SALES V01', 'RRAdsc PURCHASES-SALES V02', 'RRAdsc PURCHASES-SALES V03',
									'RRAdsc PURCHASES-SALES V04', 'RRAdsc PURCHASES-SALES V05']:
				df_mydatafileanalysed['INVOICENUMBER'] = df_mydatafileanalysed['SDCID'] + '/' + df_mydatafileanalysed['SDCRECEIPTTYPECOUNTER']
				df_mydatafileanalysed = df_mydatafileanalysed[lst_headers_RRAdatascfile]
				df_mydatafileanalysed.columns = lst_Harmonizedheaders_RRAdatascfile
		else:
			df_mydatafileanalysed = pd.DataFrame()

		return df_mydatafileanalysed

	@classmethod
	def fn_handle_specific_cases(cls, str_Filecategory, df_CustomExcelfileconverted, dic_dfs_CustomExcelfiles={}, dic_Allprocesseddatasets={}):
		# Handling SPECIFIC CASES:

		# Handling exchange of negative prices into negative quantities for Details of Ebm sales data:
		if str_Filecategory in ['EBM BO SALES DETAILS V20','EBM BO SALES DETAILS V21',]:
			str_fromcolmn = 'UNIT PRICE'
			str_tocolmn = 'QUANTITY'
			df_CustomExcelfileconverted = cls_Customfiles_Filetypehandler.fn_exchange_negativesign(df_CustomExcelfileconverted,str_fromcolmn,str_tocolmn,)

		if ("PAYROLL" in str_Filecategory) and (len(df_CustomExcelfileconverted.index)!=0):
				df_CustomExcelfileconverted = cls_Customfiles_Filetypehandler.fn_update_transactiondate_ifNA(df_CustomExcelfileconverted)

		if str_Filecategory in ['EBM DEVICE PURCHASES V21',]:
			str_criteriafield = 'CLIENT TIN'
			lst_Tin2remove = ['','Tin','tin','TIN','CLIENT TIN']
			lst_Header_invoicesummary= df_CustomExcelfileconverted.columns.tolist()
			lst_Header_invoicedetails=cls.df_Findap_newheaders['EBM DEVICE PURCHASES V21 DETAILS'].tolist()
			lst_Header_invoicedetails = [str_value for str_value in lst_Header_invoicedetails if str_value not in [None, '', 'NaN', 'NaT', 'nan'] and not (pd.isna(str_value) or pd.isnull(str_value))]
			dic_key_fields2add={'INVOICE SER NUMBER':'INVOICE SER NUMBER'}
			dic_fields2add={'PURCHASE STATUS':'PURCHASES STATUS', 'MODDT':'UPDATE DATE','SPPLRSDCID':'SUPPLIER SDC ID'}
			lst_col2keep = ['ORIGIN_FILE','ORIGIN_SHEETNAME','DATA GROUP','CATEGORY GROUP','FINANCIAL STATEMENT GROUP']
			
			df_Invoicesummary,df_Invoicedetails = cls_Customfiles_Filetypehandler.fn_get_invoicesdfs(df_CustomExcelfileconverted,str_criteriafield,lst_Tin2remove,lst_Header_invoicesummary,lst_Header_invoicedetails,lst_col2keep,dic_key_fields2add,dic_fields2add,)

			if 'EBM DEVICE PURCHASES V21' not in dic_dfs_CustomExcelfiles:
				dic_dfs_CustomExcelfiles['EBM DEVICE PURCHASES V21']= df_Invoicesummary
			else:
				dic_dfs_CustomExcelfiles['EBM DEVICE PURCHASES V21']= pd.concat([dic_dfs_CustomExcelfiles['EBM DEVICE PURCHASES V21'],df_Invoicedetails])
			dic_Allprocesseddatasets['EBM DEVICE PURCHASES V21'] = df_Invoicesummary 
			
			str_Filecategory = 'EBM DEVICE PURCHASES V21 DETAILS'
			df_CustomExcelfileconverted = df_Invoicedetails

		if str_Filecategory in ['EBM DEVICE SALES V21',]:
			str_criteriafield = 'SUPPLIER TIN'
			lst_Tin2remove = ['','Tin','tin','TIN','SUPPLIER TIN']
			lst_Header_invoicesummary= df_CustomExcelfileconverted.columns.tolist()
			lst_Header_invoicedetails=cls.df_Findap_newheaders['EBM DEVICE SALES V21 DETAILS'].tolist()
			lst_Header_invoicedetails = [str_value for str_value in lst_Header_invoicedetails if str_value not in [None, '', 'NaN', 'NaT', 'nan'] and not (pd.isna(str_value) or pd.isnull(str_value))]
			dic_key_fields2add={'INVOICE SER NUMBER':'INVOICE SER NUMBER'}
			dic_fields2add={'SALE STATUS':'SALES STATUS', 'MODDT':'UPDATE DATE','PURCHASE CODE':'PURCHASE CODE'}
			lst_col2keep = ['ORIGIN_FILE','ORIGIN_SHEETNAME','DATA GROUP','CATEGORY GROUP','FINANCIAL STATEMENT GROUP',]
			
			df_Invoicesummary,df_Invoicedetails = cls_Customfiles_Filetypehandler.fn_get_invoicesdfs(df_CustomExcelfileconverted,str_criteriafield,lst_Tin2remove,lst_Header_invoicesummary,lst_Header_invoicedetails,lst_col2keep,dic_key_fields2add,dic_fields2add,)

			if 'EBM DEVICE SALES V21' not in dic_dfs_CustomExcelfiles:
				dic_dfs_CustomExcelfiles['EBM DEVICE SALES V21']= df_Invoicesummary
			else:
				dic_dfs_CustomExcelfiles['EBM DEVICE SALES V21']= pd.concat([dic_dfs_CustomExcelfiles['EBM DEVICE SALES V21'],df_Invoicedetails])
			dic_Allprocesseddatasets['EBM DEVICE SALES V21'] = df_Invoicesummary 
			
			str_Filecategory = 'EBM DEVICE SALES V21 DETAILS'
			df_CustomExcelfileconverted = df_Invoicedetails

			# list_Processedfiles.append([str_Filename + ' [' + str_mysheetname + ']','EBM DEVICE SALES V21',len(df_Invoicesummary.index),
			# 			df_Invoicesummary['TRANSACTION DATE'].min(),df_Invoicesummary['TRANSACTION DATE'].max(),
			# 			0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,])

		# Handling Forex files: to add missing dates
		if str_Filecategory in ['BNR FOREX V01','BNR FOREX V02',]:
			lst_col2exclude = ['TRANSACTION DATE', 'ORIGIN_FILE', 'ORIGIN_SHEETNAME']
			lst_othercolumns = [str_col for str_col in df_CustomExcelfileconverted.select_dtypes(include=['object']).columns if str_col not in lst_col2exclude]
			df_CustomExcelfileconverted = cls_Customfiles_Filetypehandler.fn_insert_missingdates(df_CustomExcelfileconverted,'TRANSACTION DATE',lst_othercolumns)

		# Handling RRA datascience files: to determine category of trades
		elif 'RRAdsc' in cls.df_Categoriesparams[cls.df_Categoriesparams['CATEGORY'] == str_Filecategory]['CATEGORY GROUP'].tolist():
			df_CustomExcelfileconverted = cls_Customfiles_Filetypehandler.fn_calculate_tradescategory(df_CustomExcelfileconverted,)

		elif str_Filecategory in ['EBM SALES Z-REPORT DTL V21']:
			df_CustomExcelfileconverted = cls_Customfiles_Filetypehandler.fn_calculate_refund_status(df_CustomExcelfileconverted,)

		#Adding column of EBM REFERENCES
		if str_Filecategory in ['BANK STATEMENT ZIGAM CSS',]:
			# df_CustomExcelfileconverted['TRANSACTION AMOUNT DEBIT'] = df_CustomExcelfileconverted['TRANSACTION AMOUNT'].apply(lambda x: -1*float(str(x).strip()) if float(str(x).strip()) < 0 else 0)
			# df_CustomExcelfileconverted['TRANSACTION AMOUNT CREDIT'] = df_CustomExcelfileconverted['TRANSACTION AMOUNT'].apply(lambda x: float(str(x).strip()) if float(str(x).strip()) > 0 else 0)
			df_CustomExcelfileconverted['TRANSACTION AMOUNT DEBIT'] = df_CustomExcelfileconverted['TRANSACTION AMOUNT'].apply(lambda x: -x if x < 0 else 0)
			df_CustomExcelfileconverted['TRANSACTION AMOUNT CREDIT'] = df_CustomExcelfileconverted['TRANSACTION AMOUNT'].apply(lambda x: x if x > 0 else 0)

		# Handling the case of DHEgrp-InvoiceOcean-Sales files
		elif str_Filecategory in ['DHEgrp-Ocean-Sales','DHEgrp-Ocean-Purchases',]:
			lst_fillna_0=['INVOICE TOTAL AMOUNT without VAT','INVOICE TOTAL VAT AMOUNT','INVOICE TOTAL AMOUNT VAT included',
						'INVOICE TOTAL NET PRICE RWF','INVOICE VAT AMOUNT RWF','INVOICE TOTAL GROSS AMOUNT RWF','INVOICE PAYMENT DATE',
						'INVOICE AMOUNT PAID',]
			for str_colname in lst_fillna_0:
				df_CustomExcelfileconverted[str_colname]=df_CustomExcelfileconverted[str_colname].fillna(0)

			df_CustomExcelfileconverted.fillna(method='ffill', inplace=True)
			for col in df_CustomExcelfileconverted.columns:
				for i in range(1, len(df_CustomExcelfileconverted)):
					if pd.isna(df_CustomExcelfileconverted.loc[i, col]):
						df_CustomExcelfileconverted.loc[i, col] = df_CustomExcelfileconverted.loc[i - 1, col]
					if df_CustomExcelfileconverted.loc[i, col]=='' or df_CustomExcelfileconverted.loc[i, col]==None:
						if df_CustomExcelfileconverted.iloc[i, 0]==df_CustomExcelfileconverted.iloc[i-1, 0]:
							df_CustomExcelfileconverted.loc[i, col] = df_CustomExcelfileconverted.loc[i - 1, col]

			#Adding column of EBM REFERENCES
			if str_Filecategory in ['DHEgrp-Ocean-Sales',]:
				df_CustomExcelfileconverted['EBM REFERENCES'] = df_CustomExcelfileconverted['INVOICE NOTES'].apply(lambda x:cls_Customfiles_Filetypehandler.fn_extract_Ebmreferences(x, 'SALES','ebm'))

			elif str_Filecategory in ['DHEgrp-Ocean-Purchases',]:
				df_CustomExcelfileconverted['EBM REFERENCES'] = df_CustomExcelfileconverted['INVOICE NOTES'].apply(lambda x:cls_Customfiles_Filetypehandler.fn_extract_Ebmreferences(x, 'PURCHASES','sdc'))


		return df_CustomExcelfileconverted

	@staticmethod
	def fn_to_excel_multiple_sheets(data_dict: Dict[str, pd.DataFrame]) -> bytes:
		"""
		Create Excel file with multiple sheets
		
		Args:
			data_dict: Dictionary of {sheet_name: DataFrame}
			
		Returns:
			Excel file as bytes
		"""
		output = BytesIO()
		with pd.ExcelWriter(output, engine='openpyxl') as writer:
			for sheet_name, df in data_dict.items():
				df.to_excel(writer, index=False, sheet_name=sheet_name)
		return output.getvalue()
	
	@staticmethod
	def fn_read_uploaded_file(uploaded_file) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
		"""
		Read uploaded Excel or CSV file
		
		Args:
			uploaded_file: Streamlit UploadedFile object
			
		Returns:
			Tuple of (DataFrame, error_message)
		"""
		try:
			if uploaded_file.name.endswith('.csv'):
				df = pd.read_csv(uploaded_file)
			elif uploaded_file.name.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
				df = pd.read_excel(uploaded_file)
			else:
				return None, f"Unsupported file type: {uploaded_file.name}"
			
			return df, None
			
		except Exception as e:
			return None, f"Error reading file: {str(e)}"
	
	@staticmethod
	def fn_read_excel_multiple_sheets(uploaded_file) -> Tuple[Optional[Dict[str, pd.DataFrame]], Optional[str]]:
		"""
		Read Excel file with multiple sheets
		
		Args:
			uploaded_file: Streamlit UploadedFile object
			
		Returns:
			Tuple of (Dict of {sheet_name: DataFrame}, error_message)
		"""
		try:
			if not uploaded_file.name.endswith(('.xlsx', '.xls', '.xlsm', '.xlsb')):
				return None, "File must be an Excel file"
			
			# Read all sheets
			excel_file = pd.ExcelFile(uploaded_file)
			sheets_dict = {}
			
			for sheet_name in excel_file.sheet_names:
				sheets_dict[sheet_name] = pd.read_excel(excel_file, sheet_name=sheet_name)
			
			return sheets_dict, None
			
		except Exception as e:
			return None, f"Error reading Excel file: {str(e)}"
	
	@staticmethod
	def fn_create_download_button(label: str, data: bytes, filename: str, mime: str = None):
		"""
		Create a download button for data
		
		Args:
			label: Button label
			data: Data to download
			filename: Name of the file
			mime: MIME type
		"""
		if mime is None:
			if filename.endswith('.xlsx'):
				mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
			elif filename.endswith('.csv'):
				mime = "text/csv"
		
		return st.download_button(
			label=label,
			data=data,
			file_name=filename,
			mime=mime
		)

	# ============================================================================
	# METHODS TO ADD TO cls_Customfiles_Filetypehandler CLASS
	# ============================================================================

	@classmethod
	def get_financial_statement_groups(cls) -> Dict[str, str]:
		"""
		Get mapping of categories to financial statement groups
		
		Returns:
			Dictionary {category: financial_statement_group}
		"""
		cls.load_parameters()
		
		if cls.df_Categoriesparams.empty:
			return {}
		
		# Check if FINANCIAL STATEMENT GROUP column exists
		if 'FINANCIAL STATEMENT GROUP' not in cls.df_Categoriesparams.columns:
			print("⚠️ Warning: FINANCIAL STATEMENT GROUP column not found in Categories_sheetnames")
			return {}
		
		return dict(zip(
			cls.df_Categoriesparams['CATEGORY'],
			cls.df_Categoriesparams['FINANCIAL STATEMENT GROUP']
		))

	@classmethod
	def get_categories_by_group(cls, group_name: str) -> List[str]:
		"""
		Get all categories belonging to a financial statement group
		
		Args:
			group_name: Financial statement group name
		
		Returns:
			List of category names
		"""
		cls.load_parameters()
		
		if cls.df_Categoriesparams.empty:
			return []
		
		if 'FINANCIAL STATEMENT GROUP' not in cls.df_Categoriesparams.columns:
			return []
		
		mask = cls.df_Categoriesparams['FINANCIAL STATEMENT GROUP'] == group_name
		return cls.df_Categoriesparams[mask]['CATEGORY'].tolist()

